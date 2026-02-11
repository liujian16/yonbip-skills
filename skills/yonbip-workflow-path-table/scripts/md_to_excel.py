#!/usr/bin/env python3
"""
Convert Markdown table to YonBIP Excel format.

Usage: python md_to_excel.py <input.md> <output.xlsx> --condition-cols <N>
"""

import sys
import re
import argparse
from typing import List, Dict, Tuple


def parse_args():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(
        description='Convert Markdown table to YonBIP Excel format'
    )
    parser.add_argument('input_md', help='Input Markdown file')
    parser.add_argument('output_xlsx', help='Output Excel file')
    parser.add_argument(
        '--condition-cols',
        type=int,
        required=True,
        help='前N列为"条件"列，其余为"环节"列'
    )
    return parser.parse_args()


def parse_markdown_table(md_file: str) -> Tuple[List[str], List[Dict[str, str]]]:
    """Parse Markdown table into headers and data rows."""
    with open(md_file, 'r', encoding='utf-8') as f:
        lines = [line.strip() for line in f if line.strip()]

    # Find table boundaries
    table_lines = []
    in_table = False
    for line in lines:
        if line.startswith('|'):
            in_table = True
            table_lines.append(line)
        elif in_table:
            break

    if not table_lines:
        raise ValueError("No Markdown table found in file")

    # Parse header
    headers = [h.strip() for h in table_lines[0].split('|')[1:-1]]

    # Skip separator line (line 1)
    # Parse data rows
    data = []
    for line in table_lines[2:]:
        values = [v.strip() for v in line.split('|')[1:-1]]
        row = dict(zip(headers, values))
        data.append(row)

    return headers, data


def column_letter(col_idx: int) -> str:
    """将列索引转换为列字母 (1->A, 2->B, 27->AA)"""
    result = ""
    while col_idx > 0:
        col_idx -= 1
        result = chr(col_idx % 26 + ord('A')) + result
        col_idx //= 26
    return result


def generate_excel(
    headers: List[str],
    data: List[Dict[str, str]],
    output_file: str,
    condition_cols: int
):
    """
    Generate Excel file with YonBIP formatting.

    Args:
        headers: Column headers
        data: Data rows
        output_file: Output Excel file path
        condition_cols: Number of condition columns (from left)
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("Error: openpyxl not installed. Install with: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "流程路径表"

    # Styles
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    current_row = 1

    # ===== Step 1: Write explanation row (条件 | 环节) =====
    # 只有当 condition_cols > 0 且 < len(headers) 时才添加说明行
    if condition_cols > 0 and condition_cols < len(headers):
        # Merge "条件" cells (columns 1 to condition_cols)
        cond_end_col = column_letter(condition_cols)
        ws.merge_cells(f'A{current_row}:{cond_end_col}{current_row}')
        cell = ws.cell(row=current_row, column=1, value='条件')
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

        # Merge "环节" cells (columns condition_cols+1 to end)
        link_start_col = column_letter(condition_cols + 1)
        link_end_col = column_letter(len(headers))
        ws.merge_cells(f'{link_start_col}{current_row}:{link_end_col}{current_row}')
        cell = ws.cell(row=current_row, column=condition_cols + 1, value='环节')
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

        current_row += 1

    # ===== Step 2: Write headers =====
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    current_row += 1

    # ===== Step 3: Write data rows =====
    for row_data in data:
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, '')

            # Convert n/a to empty string
            if value == 'n/a':
                value = ''

            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

        current_row += 1

    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        col_letter = column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Set row height for header rows
    ws.row_dimensions[1].height = 25  # Explanation row
    ws.row_dimensions[2].height = 25  # Header row

    wb.save(output_file)
    print(f"Excel file generated: {output_file}")
    print(f"Total data rows: {len(data)}")


def main():
    args = parse_args()

    try:
        headers, data = parse_markdown_table(args.input_md)
        generate_excel(
            headers=headers,
            data=data,
            output_file=args.output_xlsx,
            condition_cols=args.condition_cols
        )
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
